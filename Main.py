# CODE WRITTEN BY JOSHUA LEE,
# GITHUB: scaryjoshie
# EMAIL: scaryjoshie@gmail.com
import json
import time
import glob
import os
import shutil
import configparser
from multiprocessing import Pool

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

from src.TableOperations import Table
from src.MiscUtils import location_to_cell_id, foo
from src.DuplicateShift import duplicate_shift
from src.SimpleRowFix import simple_row_fix

# gets config from config.ini
config_parser = configparser.ConfigParser()
config_parser.read("config.ini")
config_parser.sections()
config = config_parser["config"]


# color mapping
fillers = {}
# a = config
color_dict = dict(config_parser.items("COLOR_DICT"))
for color in color_dict.keys():
    temp = PatternFill(patternType="solid", fgColor=color_dict[color])
    fillers[color] = temp

# create/overwrite output dir
if config['OVERRIDE_EXISTING_OUTPUT'] and os.path.exists(config['OUTPUT_DIR']):
    shutil.rmtree(config['OUTPUT_DIR'])
else:
    AssertionError("Please Clear or Change Output Dir")
os.mkdir(config['OUTPUT_DIR'])


def process_spreadsheet(file_path):
    print(file_path)

    # Reads file into df then runs through the de-merger
    raw_df = pd.read_excel(file_path, "Table_0")
    #df = simple_row_fix(raw_df)
    df = raw_df
    #df = duplicate_shift(raw_df)

    # Creates table object and checks all values
    table = Table(df=df)
    table.check_dates()
    table.check_normals(QUOTIENT_MAX=int(config['QUOTIENT_MAX']))
    table.check_labels()

    # Loads file into openpyxl
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Table_0"]

    # Fills every cell value with what it's supposed to be
    for cell in table.df_list:
        ws[location_to_cell_id(cell.location)] = cell.value

    # Fills in all colors
    for type in table.bad_cells.keys():
        for cell in table.bad_cells[type]:
            ws[location_to_cell_id(cell.location)].fill = fillers[type]

    # Names and saves file
    name = os.path.basename(file_path)
    wb.save(os.path.join(config['OUTPUT_DIR'], name))


# Process all files in parallel
if __name__ == "__main__":
    start_time = time.perf_counter()  # for timing execution

    file_paths = glob.glob(f"{config['INPUT_DIR']}/*.xlsx")

    for i in file_paths:
        process_spreadsheet(file_path=i)


    '''
    with Pool(processes=4) as pool:  # concurrent execution
        pool.map(
            process_spreadsheet, glob.glob(f"{config['INPUT_DIR']}/*.xlsx")
        )  # inputs are a list of files in the input_dir -> passed to func
    '''
        
    

    # completion message + total time
    print(f"Spread sheet processing completed in: {time.perf_counter()-start_time}")
