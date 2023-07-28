# Imports
import glob
import pandas as pd
import os
import openpyxl

# Gets paths of every xlsx file in specified directory
FOLDER_PATH = "Input\\Completed Tables"
file_paths = glob.glob(f"{FOLDER_PATH}/*.xlsx")
file_paths = [file for file in file_paths if "all_page_firm" not in file]

# Opens every file in the path
file_list = []
for file_path in file_paths:
    # Prints file path
    print(file_path)

    # Reads the excel file into a pandas dataframe
    queries = pd.read_excel(file_path, "Queries")

    # Adds every row in the table to the dictionary with key "variable" and value "value"
    file_dict = {}
    for row in range(0, queries.shape[0]):
        variable_name = queries.loc[row]["variable"]
        value = queries.loc[row]["value"]
        file_dict[variable_name] = value
    
    # Adds the dictionary for current table to a list containing all other tables
    file_list.append(file_dict)

# Sorts file list
file_list.sort(key=lambda x: int(x["page number"]), reverse=False)


# Loads file into openpyxl
wb = openpyxl.load_workbook(f"{FOLDER_PATH}\\all_page_firm.xlsx")
ws = wb["Table_0"]

# 
start_row = 2
for file in file_list:
    ws[f"A{str(start_row)}"] = str(file["page number"])
    ws[f"B{str(start_row)}"] = file["firm name"]
    ws[f"C{str(start_row)}"] = file["stock name"]
    start_row += 1

# Names and saves file
wb.save("Output\\all_page_firm.xlsx")